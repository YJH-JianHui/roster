"""
export_utils.py
两个核心功能：
  1. export_db_to_excel(db_file, template_path) → 生成带数据的Excel（全文本格式）
  2. build_import_report(results, db_file)       → 生成导入明细Excel（分Sheet）
"""

import io
import sqlite3
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── 颜色常量（复刻模板） ─────────────────────────
C_TITLE_BG   = "1F4E79"
C_TITLE_FG   = "FFFFFF"
C_DESC_BG    = "E2EFDA"
C_DESC_FG    = "7F7F7F"
C_HDR_BG     = "D6E4F0"
C_HDR_KEY_BG = "FFC000"
C_HDR_FG     = "1F4E79"
C_DATA_EVEN  = "FFFFFF"
C_DATA_ODD   = "F5F9FF"
THIN = Side(style='thin', color="BBBBBB")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Sheet 与数据库表的映射关系 ──────────────────
# (sheet_name, table, select_sql, 列头列表, 说明)
SHEET_EXPORT_CONFIG = [
    (
        '员工主表', 'employee',
        "SELECT real_name,id_card_no,gender,birth_date,ethnicity,political_status,"
        "hometown_type,marital_status,native_place,"
        "id_card_authority,id_card_issue_date,"
        "id_card_expire_date,current_status,phone,email_personal,"
        "emergency_contact_name,emergency_contact_relation,emergency_contact_phone,"
        "domicile_province,domicile_city,domicile_district,domicile_address_detail,"
        "domicile_address_detail_extra,domicile_postal_code,domicile_hukou_type,"
        "current_province,current_city,current_district,current_address_detail,"
        "current_address_detail_extra,"
        "hire_date,tenure_base_date,pre_work_years,"
        "photo_path FROM employee ORDER BY id_card_no",
        ['姓名_real_name', '公民身份号码_id_card_no', '性别_gender', '出生年月_birth_date',
         '民族_ethnicity', '政治面貌_political_status', '户籍属性_hometown_type',
         '婚姻状况_marital_status', '籍贯_native_place',
         '发证机关_id_card_authority', '发证日期_id_card_issue_date',
         '证件到期日_id_card_expire_date', '当前状态_current_status',
         '联系方式_phone', '个人邮箱_email_personal',
         '紧急联系人_emergency_contact_name', '与本人关系_emergency_contact_relation',
         '紧急电话_emergency_contact_phone',
         '户籍省_domicile_province', '户籍市_domicile_city', '户籍区/县_domicile_district',
         '户籍详细地址_domicile_address_detail', '户籍补充地址_domicile_address_detail_extra',
         '户籍邮编_domicile_postal_code', '户口类型_domicile_hukou_type',
         '现住省_current_province', '现住市_current_city', '现住区/县_current_district',
         '现住详细地址_current_address_detail', '现住补充地址_current_address_detail_extra',
         '当前公司入职日_hire_date', '连续司龄起算日_tenure_base_date', '工前年限_pre_work_years',
         '照片路径_photo_path'],
        '【说明】所有基础信息（含证件、地址）都在此表维护。每名员工一行；公民身份号码为唯一标识。',
    ),
    (
        '任职记录', 'employment_record',
        "SELECT e.real_name,r.id_card_no,r.start_date,r.end_date,r.record_type,"
        "r.company,r.labor_relation_company,r.dept_level1,r.dept_level2,r.dept_level3,"
        "r.group_name,r.position_name,r.job_level,r.job_class,r.job_level_class,"
        "r.salary_amount,r.change_reason,"
        "r.end_reason,r.contract_expire_date,"
        "r.non_compete_signed,"
        "r.change_scope,r.change_category,r.salary_change_attr "
        "FROM employment_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.start_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no',
         '本段开始日期_start_date',
         '本段结束日期_end_date',
         '人员类型_record_type','用工公司_company',
         '劳动关系隶属_labor_relation_company','一级部门_dept_level1',
         '二级部门_dept_level2','三级部门_dept_level3','组别_group_name',
         '岗位名称_position_name','职级_job_level','职类_job_class',
         '职级职类_job_level_class','薪酬金额_salary_amount','变动原因_change_reason',
         '结束原因_end_reason',
         '合同到期日_contract_expire_date',
         '竞业限制签署_non_compete_signed',
         '变动范围_change_scope','变动类型_change_category','岗薪酬变动属性_salary_change_attr'],
        '【说明】唯一键=公民身份号码+本段开始日期；每次调岗/晋升/入职新增一行，同一人start_date不可重复。'
        '★ tenure_base_date(连续司龄起算日)=首次入职日，同一人所有行填同一值，不随调岗变化。'
        '★ 在职员工最新一行end_date留空；历史行end_date必须填写。姓名列仅供查阅，导入时忽略。',
    ),
    (
        '教育经历', 'education_record',
        "SELECT e.real_name,r.id_card_no,r.school_name,r.start_date,r.is_highest,"
        "r.degree_level,r.degree_type,r.degree_status,r.school_type,"
        "r.major,r.minor_major,"
        "r.study_duration,"
        "r.graduation_date,r.diploma_no,r.degree_cert_no "
        "FROM education_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.start_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','院校名称_school_name',
         '入学时间_start_date','是否最高学历_is_highest','学历_degree_level',
         '学位_degree_type','学习方式_degree_status','院校属性_school_type',
         '专业_major','辅修专业_minor_major',
         '学制(年)_study_duration',
         '毕业时间_graduation_date','毕业证编号_diploma_no','学位证编号_degree_cert_no'],
        '【说明】唯一键=公民身份号码+院校名称+入学时间；is_highest填1或0。',
    ),
    (
        '合同记录', 'contract_record',
        "SELECT e.real_name,r.id_card_no,r.seq,r.start_date,r.end_date,"
        "r.contract_type,r.remark "
        "FROM contract_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.seq,r.start_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','签订次数_seq',
         '开始日期_start_date','到期日期_end_date','合同类型_contract_type','备注_remark'],
        '【说明】唯一键=公民身份号码+签订次数seq+开始日期；新签合同递增seq。',
    ),
    (
        '家庭成员', 'family_member',
        "SELECT e.real_name,r.id_card_no,r.relation,r.real_name AS member_name,"
        "r.birth_date,r.political_status,r.education_level,r.work_unit,"
        "r.position,r.phone "
        "FROM family_member r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','称谓_relation',
         '成员姓名_member_name','出生年月_birth_date','政治面貌_political_status',
         '学历_education_level','工作单位_work_unit','职务_position','联系方式_phone'],
        '【说明】唯一键=公民身份号码+称谓+成员姓名；新增成员直接追加行。',
    ),
    (
        '职称职业资格', 'certificate_record',
        "SELECT e.real_name,r.id_card_no,r.cert_category,r.cert_name,r.cert_major,"
        "r.cert_level,r.issue_date,r.cert_no,r.expire_date "
        "FROM certificate_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.cert_category",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','资质类别_cert_category',
         '资质名称_cert_name','所属专业_cert_major','资质等级_cert_level',
         '取证时间_issue_date','证书编号_cert_no','到期时间_expire_date'],
        '【说明】唯一键=公民身份号码+资质名称+资质类别。',
    ),
    (
        # Sheet名：培训经历，字段名全部更新
        '培训经历', 'training_record',
        "SELECT e.real_name,r.id_card_no,r.training_project_name,r.start_date,r.end_date,"
        "r.training_type,r.training_org,r.training_hours,r.result,r.cert_obtained_flag,"
        "r.service_agreement,r.service_period "
        "FROM training_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.start_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no',
         '培训项目名称_training_project_name',
         '开始时间_start_date','结束时间_end_date','培训类型_training_type',
         '培训机构_training_org','培训学时_training_hours','考核结果_result',
         '是否获证_cert_obtained_flag',
         '是否签订服务期限协议_service_agreement','服务起止时间_service_period'],
        '【说明】唯一键=公民身份号码+培训项目名称+开始时间。',
    ),
    (
        '奖惩记录', 'reward_punishment_record',
        "SELECT e.real_name,r.id_card_no,r.record_type,r.record_date,r.reason,"
        "r.category,r.issuer "
        "FROM reward_punishment_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.record_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','类型_record_type',
         '日期_record_date','原因_reason','类别_category','签发单位_issuer'],
        '【说明】唯一键=公民身份号码+类型+日期+原因。',
    ),
    (
        # Sheet名：入职公司前工作经历，删除离职原因/证明人/证明电话，新增薪资
        '入职公司前工作经历', 'work_experience',
        "SELECT e.real_name,r.id_card_no,r.company_name,r.start_date,r.end_date,"
        "r.industry,r.company_type,r.position,r.salary "
        "FROM work_experience r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.start_date",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','工作单位_company_name',
         '开始时间_start_date','结束时间_end_date','行业_industry',
         '单位属性_company_type','职务_position','薪资_salary'],
        '【说明】唯一键=公民身份号码+工作单位+开始时间。',
    ),
    (
        '薪酬调整记录', 'salary_change_record',
        "SELECT e.real_name,r.id_card_no,r.period,r.company,r.dept,r.position,"
        "r.job_level,r.job_class,r.job_level_class,r.remark "
        "FROM salary_change_record r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no,r.period",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','时间节点_period',
         '用工公司_company','用工部门_dept','岗位_position','职级_job_level',
         '职类_job_class','职级职类_job_level_class','备注_remark'],
        '【说明】唯一键=公民身份号码+时间节点。',
    ),
    (
        '飞书账号映射', 'feishu_user_map',
        "SELECT e.real_name,r.id_card_no,r.feishu_user_id "
        "FROM feishu_user_map r LEFT JOIN employee e ON r.id_card_no=e.id_card_no "
        "ORDER BY r.id_card_no",
        ['姓名[冗余]_real_name','公民身份号码_id_card_no','飞书UserID_feishu_user_id'],
        '【说明】唯一键=公民身份号码；每个员工一行记录。',
    ),
]


def _apply_title_row(ws, title, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font      = Font(bold=True, color=C_TITLE_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _apply_desc_row(ws, desc, ncols):
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, desc)
    c.font      = Font(color=C_DESC_FG, size=9)
    c.fill      = PatternFill("solid", fgColor=C_DESC_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14


def _apply_header_row(ws, headers):
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(3, col, hdr)
        is_key = 'id_card_no' in hdr
        c.font      = Font(bold=True, color=C_HDR_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=(C_HDR_KEY_BG if is_key else C_HDR_BG))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN_BORDER
    ws.row_dimensions[3].height = 30


def _write_data_rows(ws, rows, ncols, start_row=4):
    TEXT_FMT = "@"
    for ri, row in enumerate(rows):
        excel_row = start_row + ri
        fill_color = C_DATA_EVEN if ri % 2 == 0 else C_DATA_ODD
        fill = PatternFill("solid", fgColor=fill_color)
        for ci, val in enumerate(row):
            c = ws.cell(excel_row, ci + 1)
            c.value         = str(val) if val is not None else ""
            c.number_format = TEXT_FMT
            c.fill          = fill
            c.border        = THIN_BORDER
            c.alignment     = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[excel_row].height = 16


def _auto_col_width(ws, headers, start_row=4, max_width=40, min_width=8):
    for col_idx, hdr in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        hdr_w = sum(2 if ord(c) > 127 else 1 for c in hdr.split('_')[0]) + 2
        max_w = max(hdr_w, min_width)
        for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_w = max(max_w, cell_w)
        ws.column_dimensions[col_letter].width = min(max_w + 1, max_width)


# ══════════════════════════════════════════════════════════════
#  功能1：导出数据库到Excel
# ══════════════════════════════════════════════════════════════

def export_db_to_excel(db_file: str) -> bytes:
    conn = sqlite3.connect(db_file)
    conn.row_factory = sqlite3.Row
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, table, sql, headers, desc in SHEET_EXPORT_CONFIG:
        try:
            rows = [tuple(r) for r in conn.execute(sql).fetchall()]
        except Exception:
            rows = []

        ws = wb.create_sheet(sheet_name)
        ncols = len(headers)
        _apply_title_row(ws, sheet_name, ncols)
        _apply_desc_row(ws, desc, ncols)
        _apply_header_row(ws, headers)
        _write_data_rows(ws, rows, ncols)
        _auto_col_width(ws, headers)
        ws.freeze_panes = "A4"

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  功能2：生成导入明细报告
# ══════════════════════════════════════════════════════════════

REQUIRED_FIELDS = {
    '员工主表':         ['real_name', 'id_card_no'],
    '任职记录':         ['id_card_no', 'start_date', 'record_type'],
    '教育经历':         ['id_card_no', 'school_name', 'start_date'],
    '合同记录':         ['id_card_no', 'seq', 'start_date', 'contract_type'],
    '家庭成员':         ['id_card_no', 'relation', 'real_name'],
    '职称职业资格':     ['id_card_no', 'cert_name', 'cert_category'],
    '培训经历':         ['id_card_no', 'training_project_name', 'start_date'],   # 原 training_name
    '奖惩记录':         ['id_card_no', 'record_date', 'record_type', 'reason'],
    '入职公司前工作经历':['id_card_no', 'company_name', 'start_date'],
    '薪酬调整记录':     ['id_card_no', 'period'],
    '飞书账号映射':     ['id_card_no', 'feishu_user_id'],
}

SUB_TABLE_QUERY = {
    '任职记录':         "SELECT id_card_no FROM employment_record",
    '教育经历':         "SELECT id_card_no FROM education_record",
    '合同记录':         "SELECT id_card_no FROM contract_record",
    '家庭成员':         "SELECT id_card_no FROM family_member",
    '职称职业资格':     "SELECT id_card_no FROM certificate_record",
    '培训经历':         "SELECT id_card_no FROM training_record",
    '奖惩记录':         "SELECT id_card_no FROM reward_punishment_record",
    '入职公司前工作经历':"SELECT id_card_no FROM work_experience",
    '薪酬调整记录':     "SELECT id_card_no FROM salary_change_record",
    '飞书账号映射':     "SELECT id_card_no FROM feishu_user_map",
}

RC_SUCCESS_BG = "E2EFDA"
RC_WARN_BG    = "FFF2CC"
RC_ERROR_BG   = "FFE0E0"
RC_MISS_BG    = "FFF0E6"
RC_HEADER_BG  = "1F4E79"
RC_HEADER_FG  = "FFFFFF"
RC_SUB_BG     = "E8F0FE"


def _report_header(ws, labels, col_widths=None):
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(1, ci, lbl)
        c.font      = Font(bold=True, color=RC_HEADER_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=RC_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_BORDER
    ws.row_dimensions[1].height = 22
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def _report_cell(ws, row, col, value, bg=None, bold=False, wrap=False):
    c = ws.cell(row, col, value)
    c.font      = Font(bold=bold, size=10)
    c.border    = THIN_BORDER
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def build_import_report(results: dict, db_file: str, import_excel=None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("导入汇总")
    _report_header(ws_sum,
        ["Sheet名称", "写入行数", "跳过空行", "错误行数", "主表未覆盖员工数", "状态"],
        [20, 12, 12, 12, 20, 10])

    conn = sqlite3.connect(db_file)
    all_ids = {r[0] for r in conn.execute("SELECT id_card_no FROM employee").fetchall()}

    sum_row = 2
    for sheet_name, stat in results.items():
        inserted    = stat.get('inserted', 0)
        skipped     = stat.get('skipped', 0)
        err_count   = len(stat.get('errors', []))

        uncovered = 0
        if sheet_name in SUB_TABLE_QUERY and sheet_name != '员工主表':
            try:
                sub_ids = {r[0] for r in conn.execute(SUB_TABLE_QUERY[sheet_name])}
                uncovered = len(all_ids - sub_ids)
            except Exception:
                pass

        if err_count > 0:
            status, bg = "❌ 有错误", RC_ERROR_BG
        elif uncovered > 0:
            status, bg = "⚠ 有缺失", RC_WARN_BG
        else:
            status, bg = "✓ 正常", RC_SUCCESS_BG

        for ci, val in enumerate([sheet_name, inserted, skipped, err_count, uncovered, status], 1):
            _report_cell(ws_sum, sum_row, ci, val, bg=bg)
        sum_row += 1

    conn.close()

    for sheet_name, stat in results.items():
        errors   = stat.get('errors', [])
        inserted = stat.get('inserted', 0)
        skipped  = stat.get('skipped', 0)

        ws = wb.create_sheet(sheet_name[:28])

        ws.merge_cells("A1:F1")
        c = ws.cell(1, 1, f"{sheet_name}  —  写入:{inserted}  跳过空行:{skipped}  错误:{len(errors)}")
        c.font      = Font(bold=True, size=11, color=RC_HEADER_FG)
        c.fill      = PatternFill("solid", fgColor=RC_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        cur_row = 2

        if errors:
            _report_header_inline(ws, cur_row,
                ["类型", "Excel行号", "错误原因"], [8, 12, 60])
            cur_row += 1
            for e in errors:
                _report_cell(ws, cur_row, 1, "错误行", bg=RC_ERROR_BG, bold=True)
                _report_cell(ws, cur_row, 2, e.get('row', ''), bg=RC_ERROR_BG)
                _report_cell(ws, cur_row, 3, e.get('msg', ''), bg=RC_ERROR_BG, wrap=True)
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1
        else:
            c = ws.cell(cur_row, 1, "✓ 本Sheet无错误行")
            c.font = Font(color="2E7D32", bold=True, size=10)
            cur_row += 1

        cur_row += 1

        if sheet_name != '员工主表' and sheet_name in SUB_TABLE_QUERY:
            conn2 = sqlite3.connect(db_file)
            try:
                emp_rows = conn2.execute(
                    "SELECT id_card_no, real_name FROM employee ORDER BY id_card_no"
                ).fetchall()
                sub_ids2 = {r[0] for r in conn2.execute(SUB_TABLE_QUERY[sheet_name])}
                missing_emps = [(r[0], r[1]) for r in emp_rows if r[0] not in sub_ids2]
            except Exception:
                missing_emps = []
            conn2.close()

            if missing_emps:
                _report_header_inline(ws, cur_row,
                    ["类型", "公民身份号码", "姓名", "说明"], [8, 22, 14, 40])
                cur_row += 1
                for id_no, name in missing_emps:
                    _report_cell(ws, cur_row, 1, "未覆盖", bg=RC_MISS_BG, bold=True)
                    _report_cell(ws, cur_row, 2, id_no,  bg=RC_MISS_BG)
                    _report_cell(ws, cur_row, 3, name or '', bg=RC_MISS_BG)
                    _report_cell(ws, cur_row, 4,
                        f"员工主表中存在该员工，但{sheet_name}中无记录",
                        bg=RC_MISS_BG, wrap=True)
                    ws.row_dimensions[cur_row].height = 18
                    cur_row += 1
            else:
                c = ws.cell(cur_row, 1, "✓ 所有员工在本Sheet中均有记录")
                c.font = Font(color="2E7D32", bold=True, size=10)
                cur_row += 1

        if import_excel and sheet_name in REQUIRED_FIELDS:
            missing_field_rows = _analyze_missing_fields(
                import_excel, sheet_name, REQUIRED_FIELDS[sheet_name]
            )
            if missing_field_rows:
                cur_row += 1
                _report_header_inline(ws, cur_row,
                    ["类型", "Excel行号", "公民身份号码", "缺失字段"], [8, 12, 22, 50])
                cur_row += 1
                for mfr in missing_field_rows:
                    _report_cell(ws, cur_row, 1, "缺失字段", bg=RC_WARN_BG, bold=True)
                    _report_cell(ws, cur_row, 2, mfr['row'],         bg=RC_WARN_BG)
                    _report_cell(ws, cur_row, 3, mfr['id_card_no'],  bg=RC_WARN_BG)
                    _report_cell(ws, cur_row, 4, mfr['missing'],     bg=RC_WARN_BG, wrap=True)
                    ws.row_dimensions[cur_row].height = 18
                    cur_row += 1

        ws.column_dimensions['A'].width = 10
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _report_header_inline(ws, row, labels, col_widths=None):
    for ci, lbl in enumerate(labels, 1):
        c = ws.cell(row, ci, lbl)
        c.font      = Font(bold=True, color=RC_HEADER_FG, size=10)
        c.fill      = PatternFill("solid", fgColor="37474F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_BORDER
    ws.row_dimensions[row].height = 20
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 0, w
            )


def _analyze_missing_fields(wb_import, sheet_name, required_keys):
    if sheet_name not in wb_import.sheetnames:
        return []
    ws = wb_import[sheet_name]
    rows = list(ws.iter_rows(min_row=3))
    if len(rows) < 2:
        return []

    field_keys = []
    for cell in rows[0]:
        val = str(cell.value).strip() if cell.value else ''
        field_keys.append(val.split('_', 1)[1].strip() if '_' in val else val)

    result = []
    for ri, row in enumerate(rows[1:], start=4):
        raw = {}
        for ci, cell in enumerate(row):
            if ci < len(field_keys):
                val = cell.value
                if val is not None and str(val).strip() == '':
                    val = None
                raw[field_keys[ci]] = val

        if all(v is None for v in raw.values()):
            continue

        missing = [k for k in required_keys if not raw.get(k)]
        if missing:
            result.append({
                'row':        ri,
                'id_card_no': raw.get('id_card_no') or '（无公民身份号码）',
                'missing':    '、'.join(missing),
            })

    return result