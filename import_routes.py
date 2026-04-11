import io
import os
import sqlite3
import traceback
import configparser
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, send_file, session
from openpyxl import load_workbook

from export_utils import export_db_to_excel, build_import_report

import_bp = Blueprint('import_bp', __name__)
DB_FILE      = 'data/DB.db'
PHOTO_DIR    = 'data/images'
ADMIN_CONFIG = 'admin_config.ini'


def _get_admin_password() -> str:
    """每次读取配置文件，支持不重启热更新密码"""
    cfg = configparser.ConfigParser()
    cfg.read(ADMIN_CONFIG, encoding='utf-8')
    return cfg.get('admin', 'password', fallback='')


def _is_authed() -> bool:
    return session.get('admin_authed') is True

SHEET_CONFIG = {
    '员工主表': {
        'table': 'employee',
        'unique_keys': ['id_card_no'],
        'ignore_cols': [],
    },
    '任职记录': {
        'table': 'employment_record',
        'unique_keys': ['id_card_no', 'start_date'],
        'ignore_cols': ['real_name'],
    },
    '教育经历': {
        'table': 'education_record',
        'unique_keys': ['id_card_no', 'school_name', 'start_date'],
        'ignore_cols': ['real_name'],
    },
    '地址信息': {
        'table': 'address_record',
        'unique_keys': ['id_card_no', 'address_type'],
        'ignore_cols': ['real_name'],
    },
    '合同记录': {
        'table': 'contract_record',
        'unique_keys': ['id_card_no', 'seq', 'start_date'],
        'ignore_cols': ['real_name'],
        'resolve_emp_record': True,
    },
    '家庭成员': {
        'table': 'family_member',
        'unique_keys': ['id_card_no', 'relation', 'real_name'],
        'ignore_cols': [],
        'name_is_member': True,
    },
    '职称职业资格': {
        'table': 'certificate_record',
        'unique_keys': ['id_card_no', 'cert_name', 'cert_category'],
        'ignore_cols': ['real_name'],
    },
    '培训记录': {
        'table': 'training_record',
        'unique_keys': ['id_card_no', 'training_name', 'start_date'],
        'ignore_cols': ['real_name'],
    },
    '奖惩记录': {
        'table': 'reward_punishment_record',
        'unique_keys': ['id_card_no', 'record_date', 'record_type'],
        'ignore_cols': ['real_name'],
    },
    '入职前工作经历': {
        'table': 'work_experience',
        'unique_keys': ['id_card_no', 'company_name', 'start_date'],
        'ignore_cols': ['real_name'],
    },
    '薪酬调整记录': {
        'table': 'salary_change_record',
        'unique_keys': ['id_card_no', 'period'],
        'ignore_cols': ['real_name'],
    },
    '飞书账号映射': {
        'table': 'feishu_user_map',
        'unique_keys': ['id_card_no'],
        'ignore_cols': [],
    },
}

IMPORT_ORDER = [
    '员工主表', '任职记录', '教育经历', '地址信息', '合同记录',
    '家庭成员', '职称职业资格', '培训记录', '奖惩记录', '入职前工作经历',
    '薪酬调整记录', '飞书账号映射',
]

CLEAR_ORDER = [
    '飞书账号映射', '薪酬调整记录', '入职前工作经历', '奖惩记录',
    '培训记录', '职称职业资格', '家庭成员', '合同记录',
    '地址信息', '教育经历', '任职记录', '员工主表',
]


def parse_header(row):
    keys = []
    for cell in row:
        val = str(cell.value).strip() if cell.value else ''
        keys.append(val.split('_', 1)[1].strip() if '_' in val else val)
    return keys


def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def import_sheet(conn, ws, sheet_name, cfg):
    stats = {'inserted': 0, 'skipped': 0, 'errors': []}
    rows = list(ws.iter_rows(min_row=3))
    if not rows:
        return stats

    field_keys     = parse_header(rows[0])
    data_rows      = rows[1:]
    table          = cfg['table']
    unique_keys    = cfg['unique_keys']
    ignore_cols    = set(cfg.get('ignore_cols', []))
    name_is_member = cfg.get('name_is_member', False)
    resolve_emp    = cfg.get('resolve_emp_record', False)

    for row_idx, row in enumerate(data_rows, start=4):
        raw = {}
        for col_idx, cell in enumerate(row):
            if col_idx >= len(field_keys):
                break
            key = field_keys[col_idx]
            val = cell.value
            if val is not None and str(val).strip() == '':
                val = None
            raw[key] = val

        if all(v is None for v in raw.values()):
            stats['skipped'] += 1
            continue

        try:
            id_card_no = raw.get('id_card_no')
            if not id_card_no:
                raise ValueError("缺少身份证号 id_card_no")

            if table != 'employee':
                exists = conn.execute(
                    "SELECT 1 FROM employee WHERE id_card_no=?", (id_card_no,)
                ).fetchone()
                if not exists:
                    raise ValueError(f"身份证号 {id_card_no} 在员工主表中不存在")

            record = {k: v for k, v in raw.items() if k and k not in ignore_cols}

            if name_is_member and 'member_name' in record:
                record['real_name'] = record.pop('member_name')

            if resolve_emp:
                emp_row = conn.execute(
                    "SELECT id FROM employment_record "
                    "WHERE id_card_no=? AND start_date<=? "
                    "ORDER BY start_date DESC LIMIT 1",
                    (id_card_no, record.get('start_date', ''))
                ).fetchone()
                record['employment_record_id'] = emp_row[0] if emp_row else None

            cols         = ', '.join(record.keys())
            placeholders = ', '.join('?' for _ in record)
            conn.execute(
                f"INSERT INTO {table} ({cols}) VALUES ({placeholders})",
                list(record.values())
            )
            stats['inserted'] += 1

        except Exception as e:
            stats['errors'].append({
                'row':  row_idx,
                'msg':  str(e),
            })

    return stats


# ══════════════════════════════════════════════════════════════
#  路由
# ══════════════════════════════════════════════════════════════

@import_bp.route('/api/admin/verify', methods=['POST'])
def admin_verify():
    """验证管理密码，通过后写入 session"""
    data = request.get_json(silent=True) or {}
    pwd  = data.get('password', '')
    if pwd and pwd == _get_admin_password():
        session['admin_authed'] = True
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '密码错误'}), 401


@import_bp.route('/api/admin/status', methods=['GET'])
def admin_status():
    """前端页面加载时查询当前 session 是否已通过验证"""
    return jsonify({'authed': _is_authed()})


@import_bp.route('/api/admin/logout', methods=['POST'])
def admin_logout():
    session.pop('admin_authed', None)
    return jsonify({'success': True})


def _require_admin():
    """用于在接口内部做鉴权，未通过返回 401 Response，通过返回 None"""
    if not _is_authed():
        from flask import make_response
        return make_response(jsonify({'success': False, 'message': '未授权，请先验证管理密码'}), 401)
    return None


@import_bp.route('/import')
def import_page():
    return render_template('import.html')


@import_bp.route('/api/export', methods=['GET'])
def do_export():
    """导出数据库所有数据为Excel（复刻模板格式，全文本）"""
    err = _require_admin()
    if err: return err
    try:
        data = export_db_to_excel(DB_FILE)
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            io.BytesIO(data),
            as_attachment=True,
            download_name=f'员工档案导出_{ts}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@import_bp.route('/api/import', methods=['POST'])
def do_import():
    err = _require_admin()
    if err: return err
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未收到文件'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'message': '仅支持 .xlsx 格式'}), 400

    # 读取文件内容（需要读两次：一次导入，一次生成报告）
    file_bytes = file.read()

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'文件解析失败：{e}'}), 400

    sheet_names_in_file = set(wb.sheetnames)
    sheets_to_import = [s for s in IMPORT_ORDER if s in sheet_names_in_file]
    sheets_to_clear  = [s for s in CLEAR_ORDER  if s in sheet_names_in_file]

    results     = {}
    total       = {'inserted': 0, 'skipped': 0, 'error_count': 0}
    report_bytes = None

    conn = get_db()
    try:
        conn.execute("BEGIN")

        for sheet_name in sheets_to_clear:
            table = SHEET_CONFIG[sheet_name]['table']
            conn.execute(f"DELETE FROM {table}")

        for sheet_name in sheets_to_import:
            stats = import_sheet(conn, wb[sheet_name], sheet_name, SHEET_CONFIG[sheet_name])
            results[sheet_name]   = stats
            total['inserted']    += stats['inserted']
            total['skipped']     += stats['skipped']
            total['error_count'] += len(stats['errors'])

        if total['error_count'] == 0:
            conn.execute("COMMIT")
            success = True
            message = f'全量覆盖导入成功，共写入 {total["inserted"]} 条记录。'
        else:
            conn.execute("ROLLBACK")
            success = False
            message = (f'存在 {total["error_count"]} 处错误，已全部回滚。'
                       f'请修正后重新导入。')

    except Exception as e:
        conn.execute("ROLLBACK")
        conn.close()
        return jsonify({'success': False, 'message': f'导入过程异常：{e}', 'results': {}}), 500

    # 生成导入明细报告（无论成功失败都生成）
    try:
        wb_import = load_workbook(io.BytesIO(file_bytes), data_only=True)
        report_bytes = build_import_report(results, DB_FILE, wb_import)
        import base64
        report_b64 = base64.b64encode(report_bytes).decode()
    except Exception:
        report_b64 = None

    conn.close()
    return jsonify({
        'success':    success,
        'message':    message,
        'total':      total,
        'report_b64': report_b64,   # 前端用来触发下载
        'results': {
            k: {
                'inserted': v['inserted'],
                'skipped':  v['skipped'],
                'errors':   v['errors'],
            } for k, v in results.items()
        }
    })


@import_bp.route('/api/upload-photos', methods=['POST'])
def upload_photos():
    err = _require_admin()
    if err: return err
    files = request.files.getlist('photos')
    if not files:
        return jsonify({'success': False, 'message': '未收到任何文件'}), 400

    os.makedirs(PHOTO_DIR, exist_ok=True)
    saved, replaced, errors = [], [], []
    allowed_ext = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}

    for f in files:
        filename = os.path.basename(f.filename)
        if not filename:
            continue
        ext = os.path.splitext(filename)[1].lower()
        if ext not in allowed_ext:
            errors.append({'file': filename, 'msg': f'不支持的格式 {ext}'})
            continue
        dest = os.path.join(PHOTO_DIR, filename)
        is_replace = os.path.exists(dest)
        try:
            f.save(dest)
            (replaced if is_replace else saved).append(filename)
        except Exception as e:
            errors.append({'file': filename, 'msg': str(e)})

    return jsonify({
        'success':  len(errors) == 0,
        'message':  f'新增 {len(saved)} 张，覆盖 {len(replaced)} 张，失败 {len(errors)} 张',
        'saved':    saved,
        'replaced': replaced,
        'errors':   errors,
    })
